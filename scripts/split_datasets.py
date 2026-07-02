#!/usr/bin/env python3
"""Split the master workbook (v0.5) into the two released datasets.

Usage: python scripts/split_datasets.py path/to/reg142_dataset_v0.5.xlsx
Writes data/documentary.csv and data/analytical.csv, joined on unit_id.
Unknown columns are reported and kept in the documentary dataset.
"""
import sys, csv
from openpyxl import load_workbook

DOC = ["register_id","unit_id","folio_start","folio_end","date_original","date_iso",
       "date_precision","marginal_note_raw","marginal_note_present","signatories_raw",
       "signatories_norm","signatory_count","signatories_complete","text_diplomatic",
       "transcription_status","relation_type","related_unit_id","date_check","notes"]
ANA = ["unit_id","policy_domain","decision_orientation","document_trigger","actors",
       "geography","risk_terms","fiscal_terms","institutional_terms",
       "monetary_expressions","deontic_formulas"]

def main(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    unknown = [h for h in header if h and h not in DOC and h not in ANA
               and not any(h.startswith(a + "_") for a in ANA)]
    if unknown:
        print("Unknown columns kept in documentary.csv:", unknown)
    doc_cols = [h for h in header if h in DOC or h in unknown]
    ana_cols = ["unit_id"] + [h for h in header
                              if (h in ANA and h != "unit_id")
                              or any(h.startswith(a + "_") for a in ANA if a != "unit_id")]
    def dump(cols, out):
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(cols)
            for r in rows[1:]:
                w.writerow([r[idx[c]] if c in idx and idx[c] < len(r) else "" for c in cols])
        print("written", out, f"({len(rows)-1} rows, {len(cols)} cols)")
    dump(doc_cols, "data/documentary.csv")
    dump(ana_cols, "data/analytical.csv")

if __name__ == "__main__":
    main(sys.argv[1])
